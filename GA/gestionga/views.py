from django.shortcuts import render, HttpResponseRedirect
from django.http import HttpResponse
from django.http import JsonResponse
from .models import Domaine, Programme, Profil, Droit, Habilitation, Ressource, Application, Dictionnaire, Sgci, Sgsn
from django.core.paginator import Paginator, EmptyPage
from django.shortcuts import render, redirect
from django.db.models import Q
from django.contrib.auth import login, authenticate
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth import login
from django.contrib.auth.views import LogoutView
from .forms import Form_domaine, Form_programme, Form_profil, Form_droit, Form_habilitation, Form_rchprogramme, Form_ressource, Form_ProgrammeAbsent, Form_ProgrammeGA
from .forms import Form_appli, Form_module, Form_Dictionnaire, RessourceApplicationForm, Form_filtre_ressappli
from django.shortcuts import render, redirect, get_object_or_404
from openpyxl import Workbook
from .models import Sgbe, Sgtg, Sgbf, Sgcongo, Sgcam, Sgm, Table_Module, RessourceApplication, RessourceOff
import json
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from .forms import SearchForm, SearchForm2
from django.utils import timezone
from django.db import IntegrityError 
import PyPDF2
from django.core.exceptions import ObjectDoesNotExist
from io import BytesIO
from django.urls import reverse
from PyPDF2 import PdfFileReader, PdfFileWriter
from pdfminer.high_level import extract_text
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle
from datetime import date
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from .models import ResultatRecherche 
from datetime import datetime
from .forms import ConnexionForm
from django.contrib.auth import logout
from openpyxl.styles import Border, Side
from django.http import Http404




##################### HOME ##############################################################.

def home(request):
    return render(request, 'gestionga/home.html')



############################ DOMAINE FONCTION AJOUT DOMAINE #####################################

def ajout_domaine(request):
    if request.method == 'POST':
        fm1 = Form_domaine(request.POST)
        if fm1.is_valid():
            nm = fm1.cleaned_data['nom_domaine']
            reg = Domaine(nom_domaine=nm)
            reg.save()
            fm1 = Form_domaine() #Formulaire est le nom du formulaire et fm est le petit nom donné au Formulaire

    donnees1 = Domaine.objects.all()  # Récupérer toutes les instances de l'objet User

    # Paginer mon tableau DOMAINE
    paginator = Paginator(donnees1, 20)  # 20 éléments par page
    page_number = request.GET.get('page')  # Récupérer le numéro de page actuel à partir du paramètre de requête 'page'
    page_obj = paginator.get_page(page_number)

    context = {
        'fm1': Form_domaine(),
        'donnees1': page_obj,
    }

    return render(request, 'gestionga/ajoutdomaine.html', context)


####################### SUPPRIMER DOMAINE #############################

def delete_domaine(request, id):
    if request.method == 'POST':
        supr = Domaine.objects.get(pk=id) 
        supr.delete()
        return HttpResponseRedirect('/')




########################## Modifier les donnees Domaines ###########################

def update_domaine (request, id):
    domaine =  Domaine.objects.get(id=id)

    if request.method == 'POST':
        fm1 = Form_domaine(request.POST, instance=domaine)
        if fm1.is_valid():
            fm1.save()
    else:
        fm1 = Form_domaine(instance=domaine)

    context = {
        'nom': domaine.nom_domaine,        # Récupérer le nom
        'fm1': fm1,                 # Transmettre le formulaire dans le contexte
    }

    return render(request, 'gestionga/updatedomaine.html', context)


############################# PROGRAMME AJOUTER PROGRAMME #############################################

def ajout_programme(request):
    if request.method == 'POST':
        fm2 =Form_programme(request.POST)
        if fm2.is_valid():
            fm2.save()
            # Réinitialisez le formulaire pour vider les champs
            fm2 = Form_programme()
            # Redirigez l'utilisateur vers une page de succès ou autre.
            #return redirect('ajoutprogramme')  

    # Récupérez les données de projet si le formulaire n'est pas soumis
    donnees2 = Programme.objects.all()
    
    # Paginer mon tableau USER
    paginator = Paginator(donnees2, 30)  # 20 éléments par page
    page_number = request.GET.get('page')  # Récupérer le numéro de page actuel à partir du paramètre de requête 'page'
    page_num = paginator.get_page(page_number)

    context = {
        'fm2':Form_programme,  # Ajoutez le formulaire au contexte pour qu'il soit rendu dans le modèle HTML
        'donnees2': page_num,
    }

    return render(request, 'gestionga/ajoutprogramme.html', context)


########################## SUPPRIMER PROGRAMME ###################################

def delete_programme(request, id):
    if request.method == 'POST':
        supr1 = Programme.objects.get(pk=id) 
        supr1.delete()
        return redirect('ajoutprogramme')  # Rediriger vers la page ajoutprojet après la suppression
    return render(request, 'ajoutprogramme.html')  # Afficher la page de suppression




############################# MODIFIER PROGRAMME ##############################  

def update_programme (request, id):
    programme =  Programme.objects.get(id=id)

    if request.method == 'POST':
        fm2 = Form_programme(request.POST, instance= programme)
        if fm2.is_valid():
            fm2.save()
    else:
        fm2 = Form_programme(instance=programme)

    context = {
        'nom_programme': programme.nom_programme,        # Récupérer le nom
        'description_programme': programme.description_programme,
        'domaine_programme': programme.domaine_programme,
        'fm2': fm2,                 # Transmettre le formulaire dans le contexte
    }

    return render(request, 'gestionga/updateprogramme.html', context)   



############################### PROFIL AJOUTER ##########################################

def ajout_profil(request):
    if request.method == 'POST':
        fm3 =Form_profil(request.POST)
        if fm3.is_valid():
            fm3.save()
            # Réinitialisez le formulaire pour vider les champs
            fm3 = Form_profil()
            # Redirigez l'utilisateur vers une page de succès ou autre.
            #return redirect('ajoutprofil')  

    # Récupérez les données de projet si le formulaire n'est pas soumis
    donnees3 = Profil.objects.all()
    
    # Paginer mon tableau USER
    paginator = Paginator(donnees3, 10)  # 10 éléments par page
    page_number = request.GET.get('page')  # Récupérer le numéro de page actuel à partir du paramètre de requête 'page'
    page_num1 = paginator.get_page(page_number)

    context = {
        'fm3':Form_profil,  # Ajoutez le formulaire au contexte pour qu'il soit rendu dans le modèle HTML
        'donnees3': page_num1,
    }

    return render(request, 'gestionga/ajoutprofil.html', context)   


############################## PROFIL SUPPRIMER ##################################

def delete_profil(request, id):
    if request.method == 'POST':
        supr2 = Profil.objects.get(pk=id) 
        supr2.delete()
        return redirect('ajoutprofil')  # Rediriger vers la page ajoutprofil après la suppression
    return render(request, 'ajoutprofil.html')  # Afficher la page de suppression  

 
################################# MODIFIER PROFIL ####################################


def update_profil (request, id):
    profil =  Profil.objects.get(id=id)

    if request.method == 'POST':
        fm3 = Form_profil(request.POST, instance= profil)
        if fm3.is_valid():
            fm3.save()
    else:
        fm3 = Form_profil(instance=profil)

    context = {
        'nom_profil': profil.nom_profil,        # Récupérer le nom
        'description_profil': profil.description_profil,
        'fm3': fm3,                 # Transmettre le formulaire dans le contexte
    }

    return render(request, 'gestionga/updateprofil.html', context)


################################# DROIT AJOUTER DROIT ####################################

def ajout_droit(request):
    if request.method == 'POST':
        fm4 = Form_droit(request.POST)
        if fm4.is_valid():
            fm4.save()
            # Réinitialisez le formulaire pour vider les champs
            fm4 = Form_droit()
            # Redirigez l'utilisateur vers une page de succès ou autre.
            #return redirect('ajoutdroit')

    # Récupérez les données de projet si le formulaire n'est pas soumis
    donnees4 = Droit.objects.all()

    context = {
        'fm4': Form_droit,  # Ajoutez le formulaire au contexte pour qu'il soit rendu dans le modèle HTML
        'donnees4': donnees4,  # Les données sans pagination
    }

    return render(request, 'gestionga/ajoutdroit.html', context)



################################# SUPPRIMER DROIT ####################################
    
def delete_droit(request, id):
    if request.method == 'POST':
        supr3 = Droit.objects.get(pk=id) 
        supr3.delete()
        return redirect('ajoutdroit')  # Rediriger vers la page ajoutdroit après la suppression
    return render(request, 'ajoutdroit.html')  # Afficher la page de suppression  


################################# MODIFIER DROIT ####################################    

def update_droit (request, id):
    droit =  Droit.objects.get(id=id)

    if request.method == 'POST':
        fm4 = Form_droit(request.POST, instance= droit)
        if fm4.is_valid():
            fm4.save()
    else:
        fm4 = Form_droit(instance=droit)

    context = {
        'libelle_droit': droit.libelle_droit,        # Récupérer le nom
        'symbole_droit': droit.symbole_droit,
        'fm4': fm4,                 # Transmettre le formulaire dans le contexte
    }

    return render(request, '/updatedroit.html', context)


########################### HABILITATION AJOUTER ####################################

def ajout_habilitation(request):
    if request.method == 'POST':
        fm5 = Form_habilitation(request.POST)
        if fm5.is_valid():
            
            habilitation = fm5.save(commit=False)  # Enregistrez l'habilitation sans la sauvegarder immédiatement

            # Assurez-vous que le champ domaine est correctement défini
            habilitation.domaine = habilitation.programme.domaine_programme

            habilitation.save()  # Maintenant, enregistrez l'habilitation

            # Réinitialisez le formulaire pour vider les champs
            fm5 = Form_habilitation()

    else:
        fm5 = Form_habilitation()

    # Récupérez les données d'habilitation si le formulaire n'est pas soumis
    donnees5 = Habilitation.objects.all()

    # Paginez le tableau des habilitations
    paginator = Paginator(donnees5, 15)  # 15 éléments par page
    page_number = request.GET.get('page')  # Récupérez le numéro de page actuel à partir du paramètre de requête 'page'
    page_num = paginator.get_page(page_number)

    context = {
        'fm5': fm5,  # Ajoutez le formulaire au contexte pour qu'il soit rendu dans le modèle HTML
        'donnees5': page_num,
    }

    return render(request, 'gestionga/ajouthabilitation.html', context)


############################### HABILITATION MODIFIER ######################################

def update_habilitation(request, id):
    # Récupérer l'objet Habilitation ou renvoyer une réponse 404 si l'ID n'existe pas
    habilitation = get_object_or_404(Habilitation, id=id)

    if request.method == 'POST':
        fm5 = Form_habilitation(request.POST, instance=habilitation)
        if fm5.is_valid():
            # Assurez-vous que le champ domaine est correctement défini
            habilitation.domaine = habilitation.programme.domaine_programme
            fm5.save()
            return redirect('ajouthabilitation')  # Rediriger après une mise à jour réussie

    else:
        fm5 = Form_habilitation(instance=habilitation)

    context = {
        'habilitation': habilitation,  # Récupérer l'objet Habilitation
        'fm5': fm5,  # Transmettre le formulaire dans le contexte
    }

    return render(request, 'gestionga/updatehabilitation.html', context)


#################### HABILIATATION SUPPRIMER ##########################

def delete_habilitation(request, id):
    if request.method == 'POST':
        supr4 = Habilitation.objects.get(pk=id) 
        supr4.delete()
        return redirect('ajouthabilitation')  # Rediriger vers la page ajoutdroit après la suppression
    return render(request, 'ajouthabilitation.html')  # Afficher la page de suppression  



###################### RECHERCHE (PROFIL PROGRAMME DROIT) #####################################

def faire_rechercheprogramme(request):
    form6 = Form_rchprogramme()
    profil = None  # Initialiser les variables en dehors de la condition
    droit = None
    programme = None
    results = Habilitation.objects.all()  # Obtenez tous les résultats par défaut

    if request.method == 'POST':
        form6 = Form_rchprogramme(request.POST)

        if form6.is_valid():
            profil = form6.cleaned_data['profil']
            droit = form6.cleaned_data['droit']
            programme = form6.cleaned_data['programme']

            if profil:
                results = results.filter(profil=profil)

            if droit:
                results = results.filter(droit=droit)

            if programme:
                results = results.filter(programme=programme)    
        
            results = results.values(
                'id',
                'profil__nom_profil',
                'profil__description_profil',
                'programme__nom_programme',
                'programme__description_programme',
                'domaine__nom_domaine',
                'droit__symbole_droit',
            )

            return JsonResponse({'results': list(results)})

    return render(request, 'gestionga/rechercheprogramme.html', {'form6': form6, 'profil': profil, 'droit': droit, 'programme': programme})

    
######################## telecharger Tableau excel ##################################


def exporter_en_excel(request):
    form6 = Form_rchprogramme()
    profil = None
    droit = None
    programme = None
    results = Habilitation.objects.all()

    if request.method == 'POST': 
        form6 = Form_rchprogramme(request.POST)

        if form6.is_valid():
            profil = form6.cleaned_data['profil']
            droit = form6.cleaned_data['droit']
            programme = form6.cleaned_data['programme']

            if profil:
                results = results.filter(profil=profil)

            if droit:
                results = results.filter(droit=droit)

            if programme:
                results = results.filter(programme=programme)    
        
            results = results.values(
                'id',
                'profil__nom_profil',
                'profil__description_profil',
                'programme__nom_programme',
                'programme__description_programme',
                'domaine__nom_domaine',
                'droit__symbole_droit',
            )

    # Créer un classeur Excel et ajouter une feuille de calcul
    wb = Workbook()
    ws = wb.active

    # Ajouter les en-têtes
    headers = [
        'ID',
        'Profil',
        'Description Profil',
        'Programme',
        'Description Programme',
        'Domaine',
        'Droit',
    ]
    ws.append(headers)

    # Ajouter les données
    for result in results:
        row = [
            result['id'],
            result['profil__nom_profil'],
            result['profil__description_profil'],
            result['programme__nom_programme'],
            result['programme__description_programme'],
            result['domaine__nom_domaine'],
            result['droit__symbole_droit'],
        ]
        ws.append(row)

    # Ajouter des bordures aux cellules
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    # Ajuster la largeur des colonnes en fonction du contenu
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Créer une réponse HTTP avec le contenu du classeur Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=resultats.xlsx'
    wb.save(response)

    return response

    


###################### CONNEXION #####################################

def user_login(request):
    if request.method == 'POST':
        form = ConnexionForm(request, request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username') 
            password = form.cleaned_data.get('password')
            user = authenticate(request, username=username, password=password)

            if user is not None:
                login(request, user)
                # Redirigez l'utilisateur vers la page souhaitée après la connexion
                return redirect('TotalDiagrammes') 
    else:
        form = ConnexionForm()

    return render(request, 'user_login.html', {'form': form})


###########################  FORMULAIRE RESSOURCE ##################################

def enregistrer_ressource(request):
    if request.method == 'POST':
        fm7 =Form_ressource(request.POST)
        if fm7.is_valid():
            fm7.save()
            # Réinitialisez le formulaire pour vider les champs
            fm7 = Form_ressource()
            # Redirigez l'utilisateur vers une page de succès ou autre.
            #return redirect('tableau_ressource')  

    # Récupérez les données de projet si le formulaire n'est pas soumis
    donnees7 = Ressource.objects.all()
    
    # Paginer mon tableau USER
    paginator = Paginator(donnees7, 15)  # 10 éléments par page
    page_number = request.GET.get('page')  # Récupérer le numéro de page actuel à partir du paramètre de requête 'page'
    page_num = paginator.get_page(page_number)

    context = {
        'fm7':Form_ressource,  # Ajoutez le formulaire au contexte pour qu'il soit rendu dans le modèle HTML
        'donnees7': page_num,
    }

    return render(request, 'gestionga/AjouterRessource.html', context)


####################### TABLEAU RESSOURCE #############################    

def recup_ressource(request):
    if request.method == 'POST':
        fm7 =Form_ressource(request.POST)
        if fm7.is_valid():
            fm7.save()
            # Réinitialisez le formulaire pour vider les champs
            fm7 = Form_ressource()
            # Redirigez l'utilisateur vers une page de succès ou autre.
            #return redirect('ajoutprogramme')  

    # Récupérez les données des ressources si le formulaire n'est pas soumis
    donnees7 = Ressource.objects.all()
    
    # Paginer mon tableau USER
    paginator = Paginator(donnees7, 15)  # 10 éléments par page
    page_number = request.GET.get('page')  # Récupérer le numéro de page actuel à partir du paramètre de requête 'page'
    page_num = paginator.get_page(page_number)

    context = {
        'fm7':Form_ressource,  # Ajoutez le formulaire au contexte pour qu'il soit rendu dans le modèle HTML
        'donnees7': page_num,
    }

    return render(request, 'gestionga/TableauRessource.html', context)


################## CODE PROGRAMME ABSENT ##########################

def rechercher_programmes_absents(request):
    form = Form_ProgrammeAbsent()
    profil_selectionne = None
    results = Programme.objects.all()  # Obtenez tous les programmes par défaut

    if request.method == 'POST':
        form = Form_ProgrammeAbsent(request.POST)

        if form.is_valid():
            profil_selectionne = form.cleaned_data['profil']

            # Récupérez la liste de tous les programmes
            tous_les_programmes = Programme.objects.all()

            if profil_selectionne:
                # Récupérez les programmes liés au profil sélectionné
                programmes_habilites = Habilitation.objects.filter(profil=profil_selectionne).values_list('programme', flat=True)

                # Obtenez les programmes qui ne sont pas liés au profil
                results = tous_les_programmes.exclude(id__in=programmes_habilites)

            results = results.values(
                'id',
                'nom_programme',
                'description_programme',
                'domaine_programme__nom_domaine',
            )

            return JsonResponse({'results': list(results)})

    return render(request, 'gestionga/ProgrammeAbsent.html', {'form': form, 'profil_selectionne': profil_selectionne})



###########################  AJOUT APPLICATION #####################

def ajout_appli(request):
    if request.method == 'POST':
        fm8 =Form_appli(request.POST) 
        if fm8.is_valid():
            fm8.save()
            # Réinitialisez le formulaire pour vider les champs
            fm8 = Form_programme()
            # Redirigez l'utilisateur vers une page de succès ou autre.
      

    # Récupérez les données de projet si le formulaire n'est pas soumis
    donnees8 = Application.objects.all()
    
    # Paginer mon tableau USER
    paginator = Paginator(donnees8, 10)  # 10 éléments par page
    page_number = request.GET.get('page')  # Récupérer le numéro de page actuel à partir du paramètre de requête 'page'
    page_num = paginator.get_page(page_number)

    context = {
        'fm8':Form_appli,  # Ajoutez le formulaire au contexte pour qu'il soit rendu dans le modèle HTML
        'donnees8': page_num,
    }

    return render(request, 'gestionga/ajoutapplication.html', context)


###################### FONCTION PROGRAMME GA ############################ 

def recherche_programme_GA(request):
    form10 = Form_ProgrammeGA()
    custom_results_list = []

    if request.method == 'POST':
        form10 = Form_ProgrammeGA(request.POST)
        if form10.is_valid():
            table = form10.cleaned_data['table']
            cmis = form10.cleaned_data['cmis']
            module = form10.cleaned_data['modules']

            # Définir un dictionnaire pour mapper les noms de table aux classes correspondantes
            table_mapping = {
                'Sgci': Sgci,
                'Sgsn': Sgsn,
                'Sgbe': Sgbe,
                'Sgtg': Sgtg,
                'Sgbf': Sgbf,
                'Sgcongo': Sgcongo,
                'Sgcam': Sgcam,
                'Sgm': Sgm,
            }

            # Vérifier si la table sélectionnée existe dans le mapping
            if table in table_mapping:
                results = table_mapping[table].objects.all()

                # Filtrer par CMIS
                if cmis:
                    results = results.filter(cmis=cmis)

                # Filtrer par module
                if module:
                    results = results.filter(module=module)

                # Créer la liste des résultats finaux
                custom_results_list = [
                    {
                        'module': getattr(item, 'module', ''),
                        'programme_filiale': getattr(item, 'programme_filiale', ''),
                        'intitule': getattr(item, 'intitule', ''),
                        'cmis': getattr(item, 'cmis', ''),
                    }
                    for item in results
                ]

                # Supprimer tous les enregistrements existants dans la table ResultatRecherche
                ResultatRecherche.objects.all().delete()

                # Ajouter les nouveaux résultats
                ResultatRecherche.objects.bulk_create([
                    ResultatRecherche(
                        module=getattr(item, 'module', ''),
                        programme_filiale=getattr(item, 'programme_filiale', ''),
                        intitule=getattr(item, 'intitule', ''),
                        cmis=getattr(item, 'cmis', ''),
                    )  
                    for item in results
                ])

                return JsonResponse({'custom_results_list': custom_results_list})

    else:
        form10 = Form_ProgrammeGA()

    return render(request, 'gestionga/rechprogrammega.html', {'form10': form10})

############################## RECHERCHER PROGRAMME GA PAR MOT ##################################

def recherche_mot_programme_GA(request):
    # Vous pouvez récupérer les résultats de la session
    custom_results_list = request.session.get('custom_results_list', [])
    search_form = SearchForm(request.POST)

    if request.method == 'POST' and search_form.is_valid():
        # Récupérer le champ de recherche
        search = search_form.cleaned_data['search']

        # Appliquer la recherche par mot aux résultats existants
        if search:
            results = [
                item for item in custom_results_list if
                search.lower() in item['module'].lower() or
                search.lower() in item['programme_filiale'].lower() or
                search.lower() in item['intitule'].lower() or
                search.lower() in item['cmis'].lower()
            ]

            return JsonResponse({'custom_results_list': results, 'search_form': search_form})

    return JsonResponse({'custom_results_list': custom_results_list, 'search_form': search_form})


########################### TELECHARGER EXCEL PROGRAMME GA ##############################


def telecharger_en_excel(request):
    form10 = Form_ProgrammeGA()
    custom_results_list = []

    if request.method == 'POST':
        form10 = Form_ProgrammeGA(request.POST)
        if form10.is_valid():
            table = form10.cleaned_data['table']
            cmis = form10.cleaned_data['cmis']
            module = form10.cleaned_data['modules']

            # Définir un dictionnaire pour mapper les noms de table aux classes correspondantes
            table_mapping = {
                'Sgci': Sgci,
                'Sgsn': Sgsn,
                'Sgbe': Sgbe,
                'Sgtg': Sgtg,
                'Sgbf': Sgbf,
                'Sgcongo': Sgcongo,
                'Sgcam': Sgcam,
                'Sgm': Sgm,
            }

            # Vérifier si la table sélectionnée existe dans le mapping
            if table in table_mapping:
                results = table_mapping[table].objects.all()

                # Filtrer par CMIS
                if cmis:
                    results = results.filter(cmis=cmis)

                # Filtrer par module
                if module:
                    results = results.filter(module=module)

                # Créer la liste des résultats finaux
                custom_results_list = [
                    {
                        'module': getattr(item, 'module', ''),
                        'programme_filiale': getattr(item, 'programme_filiale', ''),
                        'intitule': getattr(item, 'intitule', ''),  # Assurez-vous que intitule existe dans votre modèle
                        'cmis': getattr(item, 'cmis', ''),  # Assurez-vous que cmis existe dans votre modèle
                    }
                    for item in results
                ]

                # Générer le fichier Excel
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename=resultats.xlsx'
                workbook = Workbook()
                worksheet = workbook.active

                # Ajouter les en-têtes
                headers = ['Module', 'Programme Filiale', 'Intitule', 'CMIS']
                worksheet.append(headers)

                # Ajouter les données
                for result in custom_results_list:
                    row = [result['module'], result['programme_filiale'], result['intitule'], result['cmis']]
                    worksheet.append(row)

                # Sauvegarder le fichier Excel dans la réponse HTTP
                workbook.save(response)
                return response

    else:
        form10 = Form_ProgrammeGA()

    return render(request, 'gestionga/rechprogrammega.html', {'form10': form10})
       


##################### TABLEAU MODULE #######################################

def tableau_module(request):
    if request.method == 'POST':
        fm11 =Form_module(request.POST)
        if fm11.is_valid():
            fm11.save()
            # Réinitialisez le formulaire pour vider les champs
            fm11 = Form_module()
            # Redirigez l'utilisateur vers une page de succès ou autre.
            #return redirect('ajoutprogramme')  

    # Récupérez les données de projet si le formulaire n'est pas soumis
    donnees11 = Table_Module.objects.all()
    
    # Paginer mon tableau USER
    paginator = Paginator(donnees11, 20)  # 10 éléments par page
    page_number = request.GET.get('page')  # Récupérer le numéro de page actuel à partir du paramètre de requête 'page'
    page_num = paginator.get_page(page_number)

    context = {
        'fm11':Form_module,  # Ajoutez le formulaire au contexte pour qu'il soit rendu dans le modèle HTML
        'donnees11': page_num,
    }

    return render(request, 'gestionga/VoirModule.html', context)



################ TABLEAU DICTIONNAIRE #########################

def tableau_dictionnaire(request):
    # Initialiser le formulaire de recherche
    formR = SearchForm(request.GET)
    search = request.GET.get('search', '')

    # Récupérer tous les enregistrements
    donnees12 = Dictionnaire.objects.all()

    # Filtrer les résultats en fonction du terme de recherche
    if search:
        donnees12 = donnees12.filter(module__icontains=search) | \
                      donnees12.filter(programme__icontains=search) | \
                      donnees12.filter(intitule__icontains=search) | \
                      donnees12.filter(typ__icontains=search)

    # Paginer le tableau USER
    paginator = Paginator(donnees12, 100)  # 100 éléments par page
    page_number = request.GET.get('page')  # Récupérer le numéro de page actuel à partir du paramètre de requête 'page'
    page_num = paginator.get_page(page_number)

    context = {
        'formR': formR,  # Ajouter le formulaire au contexte pour qu'il soit rendu dans le modèle HTML
        'donnees12': page_num,
    }

    return render(request, 'gestionga/VoirDictionnaire.html', context)


################### AJOUTER RESSOURCE PAR APPLICATION ####################

def ajout_ressource_application(request):
    # Initialiser le formulaire même pour les requêtes GET
    fm15 = RessourceApplicationForm()

    if request.method == 'POST':
        fm15 = RessourceApplicationForm(request.POST)
        if fm15.is_valid():
            fm15.save()  # Maintenant, enregistrez

            # Réinitialisez le formulaire pour vider les champs
            fm15 = RessourceApplicationForm()

    # Récupérez les données d'habilitation si le formulaire n'est pas soumis
    donnees15 = RessourceApplication.objects.all()

    # Paginez le tableau des habilitations
    paginator = Paginator(donnees15, 20)  # 20 éléments par page
    page_number = request.GET.get('page')  # Récupérez le numéro de page actuel à partir du paramètre de requête 'page'
    page_num = paginator.get_page(page_number)

    context = {
        'fm15': fm15,  # Ajoutez le formulaire au contexte pour qu'il soit rendu dans le modèle HTML
        'donnees15': page_num,
    }

    return render(request, 'gestionga/RessourceApplication.html', context)


####################### FILTRER RESSOURCE PAR APPLICATION #######################

def filtre_ressource_appli(request):
    form16 = Form_filtre_ressappli()
    application = None  # Initialiser les variables en dehors de la condition
    ressource = None
    results = RessourceApplication.objects.all()  # Obtenez tous les résultats par défaut

    if request.method == 'POST': 
        form16 = Form_filtre_ressappli(request.POST) 

        if form16.is_valid():
            application = form16.cleaned_data['application']
            ressource = form16.cleaned_data['ressource']

            if application:
                results = results.filter(application=application)

            if ressource:
                results = results.filter(ressource=ressource)
        
            results = results.values(
                'id',
                'ressource__nom',
                'ressource__prenom',
                'ressource__email',
                'ressource__matricule',
                'application__application',
            )

            return JsonResponse({'results': list(results)})

    return render(request, 'gestionga/FiltreRessApp.html', {'form16': form16, 'application': application, 'ressource': ressource})

########################## EXTRACTION RESSOURCE APPLICATION #########################

def ExtractionRessApp(request):

    form16 = Form_filtre_ressappli()
    application = None  # Initialiser les variables en dehors de la condition
    ressource = None
    results = RessourceApplication.objects.all()  # Obtenez tous les résultats par défaut

    if request.method == 'POST': 
        form16 = Form_filtre_ressappli(request.POST) 

        if form16.is_valid():
            application = form16.cleaned_data['application']
            ressource = form16.cleaned_data['ressource']

            if application:
                results = results.filter(application=application)

            if ressource:
                results = results.filter(ressource=ressource)
        
            results = results.values(
                'id',
                'ressource__nom',
                'ressource__prenom',
                'ressource__email',
                'ressource__matricule',
                'application__application',
            )

            # Créer un classeur Excel et ajouter une feuille de calcul
            wb = Workbook()
            ws = wb.active

            # Ajouter les en-têtes
            headers = [
                'ID',
                'Nom',
                'Prenom',
                'Email',
                'Matricule',
                'Application',
            ]
            ws.append(headers)

            # Ajouter les données
            for result in results:
                row = [
                    result['id'],
                    result['ressource__nom'],
                    result['ressource__prenom'],
                    result['ressource__email'],
                    result['ressource__matricule'],
                    result['application__application'],
                ]
                ws.append(row)

            # Créer une réponse HTTP avec le contenu du classeur Excel
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=resultats.xlsx'
            wb.save(response)

            return response

    return render(request, 'gestionga/FiltreRessApp.html', {'form16': form16, 'application': application, 'ressource': ressource})


####################### DESACTIVER RESSOURCE ###############################

def CopierSupprimer(request, id):
    try:
        ressource = get_object_or_404(Ressource, id=id)

        # Copiez tous les champs de la ressource à supprimer dans RessourceOff
        ressource_desactivee = RessourceOff(
            nom=ressource.nom,
            prenom=ressource.prenom,
            email=ressource.email,
            matricule=ressource.matricule,
            numero_ip=ressource.numero_ip,
            departement=ressource.departement,
            service=ressource.service,
            statut=ressource.statut,
            manager=ressource.manager,
            date_debut=ressource.date_debut,
            date_fin=ressource.date_fin,
            date_desactivation=timezone.now(),
        )

        # Affectez l'id de la ressource à l'id de ressource_desactivee
        ressource_desactivee.id = ressource.id
        ressource_desactivee.save()

        # Supprimez la ressource de la table Ressource
        ressource.delete()

        return redirect('tableau_ressource')
    except Http404:
        # Gérer le cas où l'objet n'est pas trouvé (404 Not Found)
        return HttpResponse("La ressource que vous essayez de supprimer n'existe pas.")
    except Exception as e:
        # Imprimez la trace de la pile pour déboguer l'erreur
        print(e)
        message = "Avant de DESACTIVER la ressource merci de renseigner 'Date Fin'."
        return HttpResponse(message)
 

######################## AFFICHAGE RESSOURCE OFF ######################

def recup_ressource_off(request):
    # Récupérez les données de projet si le formulaire n'est pas soumis
    donnees17 = RessourceOff.objects.all()
    
    # Paginer mon tableau USER
    paginator = Paginator(donnees17, 15)  # 15 éléments par page
    page_number = request.GET.get('page')  # Récupérer le numéro de page actuel à partir du paramètre de requête 'page'
    page_num = paginator.get_page(page_number)

    context = {

        'donnees17': page_num,
    }

    return render(request, 'gestionga/RessourceOff.html', context)

  

################ MON FORMULAIRE PDF ##############################

def generer_formulaire_pdf(request, id):
    # Récupérer la ressource en fonction de l'ID
    ressource = get_object_or_404(Ressource, id=id)

    # Crée un nouveau document PDF
    formulaire_pdf = BytesIO()
    c = canvas.Canvas(formulaire_pdf, pagesize=letter)

    # Ajouter le logo en haut à gauche
    logo_path = r'C:\Users\Asus\Desktop\audit informatique\COFFRE FORT PROJET\STEP 9 PDF\logo.png'
    c.drawImage(logo_path, 10, 660, width=100, height=40)
     
    

    # Titre du formulaire
    c.setFont("Times-Roman", 12)
    title = "DEMANDE D’HABILITATIONS SUR DELTA BANK"
    title_width = c.stringWidth(title, "Times-Roman", 12)
    c.drawString((letter[0]-title_width)/2, 750, title)
    c.line((letter[0]-title_width)/2 - 10, 745, (letter[0]-title_width)/2 + title_width + 10, 745)
    c.line((letter[0]-title_width)/2 - 10, 765, (letter[0]-title_width)/2 + title_width + 10, 765)

    # Définition des champs par défaut ( C'est ici que les champs doivent etre remplis , donc c'est ici qu'on aura les champs remplis par defaut et ceux remplis avec les données de la ressource et les champs qui seront remplis avec les donnes de la ressources sont Nom , Prenoms, Email, Matricule, Service . Alors merci de voir comment resoudre cela)
    # Définition des champs par défaut
    

    fields = [
        ('Nom', ressource.nom),
        ('Prenoms', ressource.prenom),
        ('Email', ressource.email),
        ('Matricule', ressource.matricule),
        ('Numero', '..................................'),
        ('Departement', 'GA/SGABS'),  # Valeur par défaut pour 'Departement'
        ('Service', ressource.service),
        ('Fonction', 'Gestionnaire d\'Application'),  # Valeur par défaut pour 'Fonction'
        ('Manager', ressource.manager),
        ('Date', datetime.today().strftime('%d-%m-%Y')),  # Valeur par défaut pour 'Date' (aujourd'hui)
    ]

    y_position = 620
    for i in range(0, len(fields), 2):
        c.setFont("Times-Roman", 13)
        c.drawString(60, y_position, f"{fields[i][0]} : {fields[i][1]} ")
        if i + 1 < len(fields):
            c.drawString(350, y_position, f"{fields[i + 1][0]} : {fields[i + 1][1]} ")
        y_position -= 30

    # Ajouter les champs "Création", "Modification" et "Suppression" avec des cases à cocher
    creation_x = 60
    modification_x = 250
    suppression_x = 450
    y_position -= 30  # Descendre un peu plus pour aligner les champs et les cases à cocher

    c.setFont("Times-Roman", 12)
    c.drawString(creation_x, y_position, "Création : ")
    c.drawString(modification_x, y_position, "Modification : ")
    c.drawString(suppression_x, y_position, "Suppression : ")

    # Dessiner les cases à cocher
    checkbox_size = 13
    c.rect(creation_x + 80, y_position - 3, checkbox_size, checkbox_size, fill=1)
    c.rect(modification_x + 100, y_position - 3, checkbox_size, checkbox_size, fill=0)
    c.rect(suppression_x + 80, y_position - 3, checkbox_size, checkbox_size, fill=0)

    # Ajouter le champ "Profil demandé" avec un espace pour écrire du texte
    profil_demande_x = 60
    y_position -= 70  # Descendre pour le champ "Profil demandé"
    c.setFont("Times-Roman", 15)

    # Dessiner un rectangle autour du texte
    rectangle_width = 200
    rectangle_height = 50
    c.rect(profil_demande_x, y_position - 35, rectangle_width, rectangle_height)  

    # Écrire le texte à l'intérieur du rectangle
    texte_profil_demande = "Profil demandé :        CSMS"
    c.drawString(profil_demande_x + 5, y_position - 15, texte_profil_demande)

    
    # Rectangle noir - Agent à habiliter
    c.setFillColorRGB(0, 0, 0)  # Noir
    c.rect(0, 650, 700, 15, fill=True, stroke=False)
    c.setFillColorRGB(1, 1, 1)  # Blanc
    c.setFont("Helvetica", 12)
    text_agent = "AGENT A HABILITER"
    text_agent_width = c.stringWidth(text_agent, "Helvetica", 9)
    x_agent_position = (600 - text_agent_width) / 2
    c.drawString(x_agent_position, 654, text_agent)

    # Rectangle noir - Validation Manager
    c.setFillColorRGB(0, 0, 0)  # Noir
    c.rect(0, 260, 700, 15, fill=True, stroke=False)
    c.setFillColorRGB(1, 1, 1)  # Blanc
    c.setFont("Helvetica", 12)
    text_validation_manager = "VALIDATION MANAGER"
    text_validation_manager_width = c.stringWidth(text_validation_manager, "Helvetica", 9)
    x_validation_manager_position = (600 - text_validation_manager_width) / 2
    c.drawString(x_validation_manager_position, 264, text_validation_manager)

    # Tableau SIGNATURE ET MOTIF
    table_headers = ['Entete 1', 'Entete 2']
    c.setFont("Times-Roman", 15) 

    # Tableau SIGNATURE ET MOTIF 
    table_headers = ['Entete 1', 'Entete 2']
    c.setFont("Times-Roman", 15)

    # Données du tableau
    data = [
        ['MOTIF DE LA DEMANDE', 'DATE ET SIGNATURE DU MANAGER'],
        [f'Creation du compte Amplitude Bank du User {ressource.matricule}', '']
        ]
    table = Table(data, colWidths=[270, 270], rowHeights=[30, 120])

    # Style du tableau
    style = TableStyle([('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),  # Alignement vertical en haut (TOP)
                    ('WORDWRAP', (1, 1), (-1, -1)), # Retour automatique à la ligne
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),  # Alignement horizontal à gauche (LEFT)
                    ('LEADING', (0, 0), (-1, -1), 15),])  # Espacement entre les lignes
                 


    table.setStyle(style)

    # Position du tableau
    table.wrapOn(c, 10, 0)
    table.drawOn(c, 36, 100)

    # Pied de page
    c.setFont("Times-Italic", 8)
    pied_de_page = "Cette fiche d'habilitation est juste un exemple car les réelles sont confidentielles"
    c.drawCentredString(300, 15, pied_de_page)

    # Enregistre le fichier PDF
    c.save()
    # Récupérer le nom et le prénom de la ressource
    nom_ressource = ressource.nom
    prenom_ressource = ressource.prenom
    # Concaténer le nom du fichier avec le nom et le prénom de la ressource
    nom_fichier_pdf = f"{nom_ressource}_{prenom_ressource}_Amplitude_Bank.pdf"

    # Retourne le fichier PDF en réponse HTTP
    formulaire_pdf.seek(0)
    response = HttpResponse(formulaire_pdf.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{nom_fichier_pdf}"'

    return response



 
########## VUE RECHERCHE PAR MOT #############################

def recherche_par_mot(request):
    # Initialiser le formulaire de recherche
    formM = SearchForm2(request.GET)
    search = request.GET.get('search', '')  

    # Récupérer tous les enregistrements
    donnees50 = ResultatRecherche.objects.all() 

    # Filtrer les résultats en fonction du terme de recherche
    if search:
        donnees50 = donnees50.filter(module__icontains=search) | \
              donnees50.filter(programme_filiale__icontains=search) | \
              donnees50.filter(intitule__icontains=search) | \
              donnees50.filter(cmis__icontains=search)


    # Paginer le tableau USER
    paginator = Paginator(donnees50, 100)  # 100 éléments par page
    page_number = request.GET.get('page')  # Récupérer le numéro de page actuel à partir du paramètre de requête 'page'
    page_num = paginator.get_page(page_number)

    context = {
        'formM': formM,  # Ajouter le formulaire au contexte pour qu'il soit rendu dans le modèle HTML
        'donnees50': page_num,
    }

    return render(request, 'gestionga/recherche_par_mot.html', context)



####################### UPDATE RESSOURCE ##################################

def update_ressource(request, id):
    
    ressource = get_object_or_404(Ressource, id=id)

    if request.method == 'POST':
        fm7 = Form_ressource(request.POST, instance=ressource)
        if fm7.is_valid():
            
            fm7.save()
            return redirect('ajouterressource')  
    else:
        fm7 = Form_ressource(instance=ressource)

    context = {
        'ressource': ressource, 
        'fm7': fm7,  
    }

    return render(request, 'gestionga/UpdateRessource.html', context)




################## UPDATE APPLICATION ###################################

def update_application(request, id):
    
    application = get_object_or_404(Application, id=id)

    if request.method == 'POST':
        fm8 = Form_appli(request.POST, instance=application) 
        if fm8.is_valid():
            
            fm8.save()
            return redirect('ajoutapplication')  
    else:
        fm8 = Form_appli(instance=application)

    context = {
        'application': application, 
        'fm8': fm8,  
    }

    return render(request, 'gestionga/UpdateApplication.html', context)


##################### DECONNEXION ################################

def user_logout(request):
    logout(request)
    return redirect('connexion')


######################## UTILISATEUR CONNECTE #####################

from django.contrib.auth.decorators import login_required
from django.shortcuts import render

@login_required
def user_profile(request): 
    user = request.user
    context = {
        'nom': user.last_name,
        'prenom': user.first_name,
        'email': user.email,
    }
    return render(request, 'gestionga/MonCompte.html', context)


##################### CHANGER LE MDP ################################



############## DIAGRAMMES ###########################################

from django.shortcuts import render
from plotly.subplots import make_subplots
import plotly.graph_objects as go

def example_diagramme(request):
    # Les données pour le premier diagramme (Année 2020)
    labels1 = ['Janvier', 'Février', 'Mars']
    values1 = [10, 25, 65]

    # Les données pour le deuxième diagramme (Année 2021)
    labels2 = ['Janvier', 'Février', 'Mars']
    values2 = [25, 45, 30]

    # Les données pour le troisième diagramme (Année 2022)
    labels3 = ['Janvier', 'Février', 'Mars']
    values3 = [15, 35, 50]

    # Les données pour le quatrième diagramme (Année 2023)
    labels4 = ['Janvier', 'Février', 'Mars']
    values4 = [20, 0, 80]

    # Couleurs personnalisées pour chaque tranche
    marker_colors1 = ['#FF9999', '#66B2FF', '#99FF99']
    marker_colors2 = ['#FFCC99', '#C2C2F0', '#FFD700']
    marker_colors3 = ['#FFD700', '#FF6347', '#7B68EE']
    marker_colors4 = ['#87CEEB', '#FFD700', '#32CD32']

    # Créer le sous-diagramme avec 1 ligne et 4 colonnes
    fig = make_subplots(1, 4, specs=[[{'type':'pie'}, {'type':'pie'}, {'type':'pie'}, {'type':'pie'}]])

    # Ajouter le premier trace du diagramme (Année 2020) avec des couleurs personnalisées
    fig.add_trace(go.Pie(labels=labels1, values=values1, textinfo='label+percent', marker=dict(colors=marker_colors1)), 1, 1)

    # Ajouter le deuxième trace du diagramme (Année 2021) avec des couleurs personnalisées
    fig.add_trace(go.Pie(labels=labels2, values=values2, textinfo='label+percent', marker=dict(colors=marker_colors2)), 1, 2)

    # Ajouter le troisième trace du diagramme (Année 2022) avec des couleurs personnalisées
    fig.add_trace(go.Pie(labels=labels3, values=values3, textinfo='label+percent', marker=dict(colors=marker_colors3)), 1, 3)

    # Ajouter le quatrième trace du diagramme (Année 2023) avec des couleurs personnalisées
    fig.add_trace(go.Pie(labels=labels4, values=values4, textinfo='label+percent', marker=dict(colors=marker_colors4)), 1, 4)

    # Mettre à jour la mise en page du diagramme
    fig.update_layout(title_text='Pourcentage des différents Mois pour chaque Année',
                      showlegend=False,  # Désactive la légende des couleurs
                      height=630,  # Ajuste la hauteur du cadre (background)
                      width=1200,  # Ajuste la largeur du cadre (background)
                      # Ajuste la distance entre les diagrammes horizontalement
                      xaxis=dict(domain=[0, 0.2]),
                      xaxis2=dict(domain=[0.3, 0.5]),
                      xaxis3=dict(domain=[0.6, 0.8]),
                      xaxis4=dict(domain=[0.9, 1.0])
    )

    # Augmenter la taille des diagrammes de 30%
    fig.update_layout(
        autosize=False,
        margin=dict(l=20, r=20, t=5, b=10),  # Ajuster les marges pour le texte et les axes
        height=1.3 * fig.layout.height  # Augmenter la hauteur des diagrammes de 30%
    )

    # Ajouter les annotations pour les années avec un léger déplacement vers le bas
    fig.update_layout(
        annotations=[
            dict(text="Année 2020", x=0.1, y=0.1, font=dict(size=16, color='black')),
            dict(text="Année 2021", x=0.4, y=0.1, font=dict(size=16, color='black')),
            dict(text="Année 2022", x=0.65, y=0.1, font=dict(size=16, color='black')),
            dict(text="Année 2023", x=0.95, y=0.1, font=dict(size=16, color='black')),
        ]
    )

    # Convertir le diagramme en HTML
    chart_html = fig.to_html(full_html=False)

    # Passer le HTML du diagramme à la template
    context = {'chart_html': chart_html}
    return render(request, 'gestionga/Diagramme.html', context)



################### DIAGRAMME PROGRAMME METIER ######################################

from django.shortcuts import render
from plotly.subplots import make_subplots
import plotly.graph_objects as go
from django.db import models
from .models import Habilitation, Profil, Programme, Droit

def DiagramMetier(request):
    # Liste des profils
    profils = Profil.objects.all()

    # Créer le sous-diagramme avec 1 ligne et len(profils) colonnes
    fig = make_subplots(1, len(profils), specs=[[{'type': 'pie'}] * len(profils)], subplot_titles=[profil.nom_profil for profil in profils])

    for i, profil in enumerate(profils, start=1):
        # Récupérer les habilitations pour le profil actuel
        habilitations = Habilitation.objects.filter(profil=profil)

        # Compter le nombre de droits pour chaque type
        droits_count = habilitations.values('droit__symbole_droit').annotate(count=models.Count('droit__symbole_droit'))

        # Labels et valeurs pour le diagramme
        labels = [entry['droit__symbole_droit'] for entry in droits_count]
        values = [entry['count'] for entry in droits_count]

        # Ajouter le trace du diagramme pour le profil actuel
        fig.add_trace(go.Pie(labels=labels, values=values, textinfo='label+percent'), 1, i)

    # Mettre à jour la mise en page du diagramme
    fig.update_layout(title_text='Répartition des droits par programmes des Profils Metier',
                      showlegend=True,
                      height=1.12 * 500,  # Ajuster la hauteur du cadre (background)
                      width=1.12 * 1200  # Ajuster la largeur du cadre (background)
                      )

    # Convertir le diagramme en HTML
    chart_html = fig.to_html(full_html=False)

    # Passer le HTML du diagramme à la template
    context = {'chart_html': chart_html, 'profils': profils}
    return render(request, 'gestionga/StatProgramMetier.html', context)



########################### DIAGRAMME PROGRAMME FILIALE ########################

from django.shortcuts import render
from plotly.subplots import make_subplots
import plotly.graph_objects as go
from django.db import models
from .models import Sgci, Sgsn, Sgbe, Sgtg, Sgbf, Sgcongo, Sgcam, Sgm

def DiagramFiliale(request):
    # Liste des modèles de programmes
    program_models = [Sgci, Sgsn, Sgbe, Sgtg, Sgbf, Sgcongo, Sgcam, Sgm]

    # Diviser les modèles en deux groupes
    half_len = len(program_models) // 2
    first_group = program_models[:half_len]
    second_group = program_models[half_len:]

    # Créer le sous-diagramme avec 2 lignes et chaque ligne a la moitié des modèles
    fig = make_subplots(2, half_len, specs=[[{'type': 'pie'}] * half_len] * 2, 
                        subplot_titles=[model.__name__ for model in program_models])

    chart_html_models = []  # Liste pour stocker les HTML de chaque modèle

    for i, program_model in enumerate(first_group, start=1):
        # Récupérer les données du modèle de programme actuel
        data = program_model.objects.values('cmis').annotate(count=models.Count('cmis'))

        # Labels et valeurs pour le diagramme
        labels = [entry['cmis'] for entry in data]
        values = [entry['count'] for entry in data]

        # Ajouter le trace du diagramme pour le modèle de programme actuel
        fig.add_trace(go.Pie(labels=labels, values=values, textinfo='label+percent'), 1, i)

    for i, program_model in enumerate(second_group, start=1):
        # Récupérer les données du modèle de programme actuel
        data = program_model.objects.values('cmis').annotate(count=models.Count('cmis'))

        # Labels et valeurs pour le diagramme
        labels = [entry['cmis'] for entry in data]
        values = [entry['count'] for entry in data]

        # Ajouter le trace du diagramme pour le modèle de programme actuel
        fig.add_trace(go.Pie(labels=labels, values=values, textinfo='label+percent'), 2, i)

    # Mettre à jour la mise en page du diagramme
    fig.update_layout(
                      showlegend=True,
                      height=0.78 * 760,  # Ajuster la hauteur du cadre (background)
                      width=1.1 * 1200 # Ajuster la largeur du cadre (background)
                      )

    # Convertir le diagramme en HTML
    chart_html_model = fig.to_html(full_html=False)

    # Passer la liste de HTML des diagrammes à la template
    context = {'chart_html_model': chart_html_model, 'program_models': program_models}
    return render(request, 'gestionga/diagramme_filiale.html', context)


####################### MISE A JOUR DONNEES METIER ###############################

def MAJ(request):
    return render(request, 'gestionga/MajDonneeMetier.html')


####################### LES DIAGRAMMES ##########################################

from django.shortcuts import render

def LesDiagrammes(request):
    # Appeler la première vue (DiagramMetier)
    context_metier = CopieDiagramMetier(request)

    # Appeler la deuxième vue (DiagramFiliale)
    context_filiale = CopieDiagramFiliale(request)

    # Extraire les données pertinentes des contextes
    chart_html_metier = context_metier.get('chart_html_metier', '')
    profils = context_metier.get('profils', [])

    chart_html_model = context_filiale.get('chart_html_model', '')
    program_models = context_filiale.get('program_models', [])

    # Combiner les contextes pour les deux vues
    context = {
        'chart_html_metier': chart_html_metier,
        'profils': profils,
        'chart_html_model': chart_html_model,
        'program_models': program_models,
    }

    # Passer le contexte à la template et rendre la réponse
    return render(request, 'gestionga/LesDiagrammes.html', context)


################### COPIE DIAGRAMME METIER #######################

from django.shortcuts import render
from plotly.subplots import make_subplots
import plotly.graph_objects as go
from django.db import models
from .models import Habilitation, Profil, Programme, Droit

def CopieDiagramMetier(request):
    profils = Profil.objects.all()
    fig = make_subplots(1, len(profils), specs=[[{'type': 'pie'}] * len(profils)], subplot_titles=[profil.nom_profil for profil in profils])

    for i, profil in enumerate(profils, start=1):
        habilitations = Habilitation.objects.filter(profil=profil)
        droits_count = habilitations.values('droit__symbole_droit').annotate(count=models.Count('droit__symbole_droit'))
        labels = [entry['droit__symbole_droit'] for entry in droits_count]
        values = [entry['count'] for entry in droits_count]
        fig.add_trace(go.Pie(labels=labels, values=values, textinfo='label+percent'), 1, i)

    fig.update_layout(title_text='Répartition des droits par programmes des Profils Metier',
                       showlegend=True,
                       height=1.12 * 500,
                       width=1.12 * 1200)
    chart_html = fig.to_html(full_html=False)

    context = {'chart_html_metier': chart_html, 'profils': profils}
    return context


######################## COPIE DIAGRAMME FILIALE ########################

from django.shortcuts import render
from plotly.subplots import make_subplots
import plotly.graph_objects as go
from django.db import models
from .models import Sgci, Sgsn, Sgbe, Sgtg, Sgbf, Sgcongo, Sgcam, Sgm

def CopieDiagramFiliale(request):
    # Liste des modèles de programmes
    program_models = [Sgci, Sgsn, Sgbe, Sgtg, Sgbf, Sgcongo, Sgcam, Sgm]

    # Diviser les modèles en deux groupes
    half_len = len(program_models) // 2
    first_group = program_models[:half_len]
    second_group = program_models[half_len:]

    # Créer le sous-diagramme avec 2 lignes et chaque ligne a la moitié des modèles
    fig = make_subplots(2, half_len, specs=[[{'type': 'pie'}] * half_len] * 2, 
                        subplot_titles=[model.__name__ for model in program_models])

    chart_html_models = []  # Liste pour stocker les HTML de chaque modèle

    for i, program_model in enumerate(first_group, start=1):
        # Récupérer les données du modèle de programme actuel
        data = program_model.objects.values('cmis').annotate(count=models.Count('cmis'))

        # Labels et valeurs pour le diagramme
        labels = [entry['cmis'] for entry in data]
        values = [entry['count'] for entry in data]

        # Ajouter le trace du diagramme pour le modèle de programme actuel
        fig.add_trace(go.Pie(labels=labels, values=values, textinfo='label+percent'), 1, i)

    for i, program_model in enumerate(second_group, start=1):
        # Récupérer les données du modèle de programme actuel
        data = program_model.objects.values('cmis').annotate(count=models.Count('cmis'))

        # Labels et valeurs pour le diagramme
        labels = [entry['cmis'] for entry in data]
        values = [entry['count'] for entry in data]

        # Ajouter le trace du diagramme pour le modèle de programme actuel
        fig.add_trace(go.Pie(labels=labels, values=values, textinfo='label+percent'), 2, i)

    # Mettre à jour la mise en page du diagramme
    fig.update_layout(
                      showlegend=True,
                      height=0.78 * 760,  # Ajuster la hauteur du cadre (background)
                      width=1.1 * 1200 # Ajuster la largeur du cadre (background)
                      )

    # Convertir le diagramme en HTML
    chart_html_model = fig.to_html(full_html=False)

    # Passer la liste de HTML des diagrammes à la template
    context = {'chart_html_model': chart_html_model, 'program_models': program_models}
    return context

